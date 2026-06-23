"""WBSData를 JSON / Excel / draw.io 형식으로 내보내기.

PNG 출력의 역방향 흐름(이미지/엑셀/PDF -> Claude 파싱 -> 구조화 데이터)으로 얻은
WBSData를 다른 포맷으로 재출력할 때 사용한다.
"""
from __future__ import annotations

import base64
import io
import json

import openpyxl
from openpyxl.styles import Font, PatternFill

import renderer
from renderer import WBSData

EXCEL_HEADERS = ["파트", "파트색상", "소분류", "작업", "시작일", "종료일"]


def export_json(data: WBSData, path: str) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(renderer.to_dict(data), f, ensure_ascii=False, indent=2)


def export_excel(data: WBSData, path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WBS"
    ws.append(EXCEL_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for part in data.parts:
        part_fill = PatternFill(start_color=part.color.lstrip("#"), fill_type="solid")
        for sg in part.subgroups:
            for task in sg.tasks:
                ws.append(
                    [
                        part.name,
                        part.color,
                        sg.name,
                        task.name,
                        task.start_date.isoformat(),
                        task.end_date.isoformat(),
                    ]
                )
                ws.cell(row=ws.max_row, column=1).fill = part_fill
                ws.cell(row=ws.max_row, column=2).fill = part_fill

    for col, width in zip("ABCDEF", (18, 12, 18, 30, 14, 14)):
        ws.column_dimensions[col].width = width

    wb.save(path)


def export_drawio(data: WBSData, path: str) -> None:
    xml = renderer.GanttRenderer(data).to_drawio_xml()
    with open(path, "w", encoding="utf-8") as f:
        f.write(xml)


def export_png(data: WBSData, path: str) -> None:
    renderer.render_to_png(data, path)


def export_html(data: WBSData, path: str) -> None:
    """렌더링한 PNG를 base64로 인라인 임베드한 단독 HTML 파일로 내보낸다.
    브라우저에서 PNG 파일 없이 바로 열어볼 수 있다."""
    img = renderer.render_to_image(data)
    buf = io.BytesIO()
    img.save(buf, "PNG")
    b64 = base64.standard_b64encode(buf.getvalue()).decode("ascii")
    width, height = img.size

    html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8">
<title>WBS 간트차트</title>
<style>
  body {{ margin: 0; background: #ffffff; }}
  img {{ display: block; width: {width}px; height: {height}px; }}
</style>
</head>
<body>
<img src="data:image/png;base64,{b64}" alt="WBS 간트차트">
</body>
</html>
"""
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
