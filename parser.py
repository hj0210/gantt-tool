"""WBS 입력 파일(Excel/JSON/이미지/PDF)을 LLM(Claude/GPT/Gemini 중 선택)으로 파싱하여
간트차트 렌더링에 필요한 구조화된 데이터(part > subgroup > task)로 변환한다.
"""
import json
import os

import openpyxl

import config
import llm_providers
import settings

SYSTEM_PROMPT = """\
너는 WBS(Work Breakdown Structure) 문서를 분석해 간트차트용 JSON으로 변환하는 전문가다.
입력은 엑셀 표, 텍스트, 또는 이미지(스크린샷)일 수 있으며 포맷이 제각각이다.
다음 규칙에 따라 구조를 추론하라:

1. 데이터를 "파트(Part) > 소분류(Subgroup) > 작업(Task)" 3단계 계층으로 정리한다.
   - 소분류가 명확히 없는 경우, 파트 바로 아래에 작업을 두기 위해
     소분류 이름을 작업명과 동일하게 1개씩 만들거나, 의미 단위로 합리적으로 묶어라.
   - 어떤 컬럼을 파트로 쓸지 고를 때, 거의 모든 행에서 값이 동일한 컬럼(예: 프로젝트 전체명을
     반복해서 적어둔 컬럼)은 파트로 쓰지 마라. 그런 컬럼은 무시하고, 행마다 값이 실제로
     달라지는(=의미 있게 그룹을 나누는) 컬럼을 파트로 선택한다. 파트는 보통 5개 내외로
     색깔 구분이 의미 있는 수준이어야 하며, 파트가 1개뿐이라면 컬럼을 잘못 고른 것이다.
   - 절대 누락·요약·생략하지 마라. 입력에 들어있는 작업(최하위 항목) 행은 단 하나도
     빠짐없이 전부 포함해야 한다. 행이 많다고 대표적인 일부만 추출하거나 비슷한 행을
     하나로 합치면 안 된다. 입력 행 개수와 출력 task 개수가 최대한 일치해야 한다.
2. 각 작업은 시작일(start_date)과 종료일(end_date)을 YYYY-MM-DD 형식으로 가진다.
   - 날짜가 누락되었거나 모호하면 주변 정보(주차, 월, 순서)를 근거로 합리적으로 추론한다.
   - 날짜를 절대 추론할 수 없다면 프로젝트 시작일을 기본값으로 사용한다.
3. 프로젝트 전체 시작일(project_start)과 종료일(project_end)을 모든 작업의 최소/최대 날짜로 계산한다.
4. 각 파트에 color 필드를 부여한다. 입력에 색상 정보가 없으면 다음 팔레트를 순서대로 배정한다:
   ["#6C5CE7", "#E17055", "#0984E3", "#E84393", "#00B894", "#FDCB6E", "#A29BFE"]

절대 "분석할 수 없다", "불가능하다" 같은 거부 응답을 하지 마라. 이미지나 텍스트가 흐릿하거나
일부만 보이더라도 보이는 정보로 최선을 다해 추론하여 항상 결과를 만들어낸다.
사과 문구나 설명 문장을 절대 앞에 붙이지 말고, 반드시 아래 JSON 스키마로만 응답하라.
마크다운 코드블록(```)도 쓰지 말고 순수 JSON 객체 하나만 출력한다.

{
  "project_start": "YYYY-MM-DD",
  "project_end": "YYYY-MM-DD",
  "parts": [
    {
      "name": "파트명",
      "color": "#6C5CE7",
      "subgroups": [
        {
          "name": "소분류명",
          "tasks": [
            {"name": "작업명", "start_date": "YYYY-MM-DD", "end_date": "YYYY-MM-DD"}
          ]
        }
      ]
    }
  ]
}
"""


class WBSParseError(Exception):
    pass


_PROVIDER_ENV_VARS = {
    "anthropic": "ANTHROPIC_API_KEY",
    "openai": "OPENAI_API_KEY",
    "google": "GOOGLE_API_KEY",
}


def get_active_provider() -> str:
    return settings.get_active_provider()


def get_api_key(provider: str | None = None) -> str | None:
    """환경변수 우선, 없으면 로컬 설정 파일에서 저장된 키를 읽는다."""
    provider = provider or get_active_provider()
    env_var = _PROVIDER_ENV_VARS.get(provider)
    return (env_var and os.environ.get(env_var)) or settings.load_api_key(provider)


def validate_api_key(api_key: str, provider: str | None = None) -> tuple[bool, str]:
    """최소 토큰으로 더미 호출을 보내 키 유효성을 검증한다.
    (True, "") 또는 (False, 에러메시지) 반환."""
    provider = provider or get_active_provider()
    return llm_providers.validate_key(provider, api_key)


def _extract_json_candidate(text: str) -> str:
    """LLM 응답이 항상 순수 JSON으로만 오지는 않는다 (예: 사과 멘트 뒤에 ```json 블록을
    덧붙이는 경우). 응답 어디에 있든 JSON 블록/객체를 찾아 추출한다."""
    import re

    t = text.strip()

    fence_match = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", t, re.DOTALL | re.IGNORECASE)
    if fence_match:
        return fence_match.group(1)

    first = t.find("{")
    last = t.rfind("}")
    if first != -1 and last != -1 and last > first:
        return t[first : last + 1]

    return t


def _call_llm_text(content_text: str) -> dict:
    provider = get_active_provider()
    api_key = get_api_key(provider)
    if not api_key:
        raise WBSParseError(
            f"{config.PROVIDERS.get(provider, {}).get('label', provider)} API 키가 설정되지 않았습니다. "
            "앱에서 API 키를 입력해주세요."
        )
    raw = llm_providers.text_completion(provider, api_key, SYSTEM_PROMPT, content_text)
    return _parse_json_response(raw)


def _call_llm_vision(image_bytes: bytes, media_type: str) -> dict:
    provider = get_active_provider()
    if not config.PROVIDERS.get(provider, {}).get("vision"):
        raise WBSParseError(f"{provider} 프로바이더는 이미지(Vision) 파싱을 지원하지 않습니다.")
    api_key = get_api_key(provider)
    if not api_key:
        raise WBSParseError(
            f"{config.PROVIDERS.get(provider, {}).get('label', provider)} API 키가 설정되지 않았습니다. "
            "앱에서 API 키를 입력해주세요."
        )
    instruction = "이 이미지에 있는 WBS/일정표를 분석해서 지정된 JSON 스키마로 변환해줘."
    raw = llm_providers.vision_completion(provider, api_key, SYSTEM_PROMPT, image_bytes, media_type, instruction)
    return _parse_json_response(raw)


def _parse_json_response(raw: str) -> dict:
    try:
        return json.loads(_extract_json_candidate(raw))
    except json.JSONDecodeError as e:
        raise WBSParseError(
            f"LLM 응답을 JSON으로 해석할 수 없습니다: {e}\n"
            f"모델 응답: {raw[:500]}"
        )


# ---- 입력 포맷별 전처리 ----

SHEET_SELECT_PROMPT = """\
너는 엑셀 워크북에서 WBS(작업분류체계) 일정표가 들어있는 시트를 찾는 전문가다.
여러 시트의 이름과 앞부분 내용 미리보기가 주어진다.
표지, 변경이력, 진척현황 요약, 휴일 목록 같은 시트가 아니라
파트/작업/시작일/종료일 같은 일정 데이터가 실제로 들어있는 시트를 골라라.
동일한 내용의 시트가 여러 개(예: "...복사본")라면 그중 하나만 고른다.
다른 설명 없이 정확한 시트 이름만 한 줄로 출력한다.
"""


def _sheet_preview(ws, max_rows: int = 12, max_cols: int = 20) -> str:
    lines = [f"[시트: {ws.title}] (크기: {ws.dimensions})"]
    count = 0
    for row in ws.iter_rows(values_only=True):
        cells = [str(c) if c is not None else "" for c in row[:max_cols]]
        if any(cell.strip() for cell in cells):
            lines.append("\t".join(cells))
            count += 1
        if count >= max_rows:
            break
    return "\n".join(lines)


def _select_wbs_sheet(wb) -> str:
    """시트가 여러 개면 LLM에게 미리보기를 보여주고 실제 WBS 일정표 시트를 고르게 한다."""
    sheet_names = wb.sheetnames
    if len(sheet_names) == 1:
        return sheet_names[0]

    previews = "\n\n".join(_sheet_preview(wb[name]) for name in sheet_names)
    provider = get_active_provider()
    api_key = get_api_key(provider)
    if not api_key:
        return sheet_names[0]

    try:
        raw = llm_providers.text_completion(provider, api_key, SHEET_SELECT_PROMPT, previews)
    except Exception:
        return sheet_names[0]

    chosen = raw.strip().strip('"').strip("'")
    if chosen in sheet_names:
        return chosen
    for name in sheet_names:
        if name in chosen or chosen in name:
            return name
    return sheet_names[0]


# ---- 결정적(LLM 미사용) 엑셀 테이블 추출 ----
#
# WBS 엑셀이 "시작일/종료일" 같은 명시적 날짜 컬럼과, 그 왼쪽에 좌->우로
# 넓은 범주->좁은 범주 순서의 계층 컬럼(보통 병합 셀로 들여쓰기 표현)을 갖는
# 흔한 템플릿이면, LLM 없이 셀 좌표만으로 100% 정확하게 파트>소분류>작업을 복원한다.
# (LLM은 큰 표를 만나면 일부 행을 요약/생략하는 경향이 있어 완전성을 보장 못 함)
import datetime as _dt
import re as _re

_START_KEYWORDS = ("시작", "start")
_END_KEYWORDS = ("종료", "완료", "end", "finish")
_INDEX_COL_RE = _re.compile(r"^(no|#|순번|번호|seq|index)$", _re.IGNORECASE)


def _cell_to_date_str(value) -> str | None:
    if isinstance(value, _dt.datetime):
        return value.date().isoformat()
    if isinstance(value, _dt.date):
        return value.isoformat()
    if isinstance(value, str):
        m = _re.match(r"^(\d{4})[-./](\d{1,2})[-./](\d{1,2})", value.strip())
        if m:
            y, mo, d = (int(g) for g in m.groups())
            try:
                return _dt.date(y, mo, d).isoformat()
            except ValueError:
                return None
    return None


def _find_header_row(ws, max_scan_rows: int = 30):
    """시작/종료 날짜 컬럼이 모두 있는 첫 행을 헤더로 본다.
    반환: (header_row_idx, start_col_idx, end_col_idx, {col_idx: header_text}) 또는 None."""
    for row_idx in range(1, max_scan_rows + 1):
        texts = {}
        start_idx = end_idx = None
        for col_idx in range(1, ws.max_column + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            text = str(v).strip()
            if not text:
                continue
            texts[col_idx] = text
            low = text.lower()
            if start_idx is None and any(k in text or k in low for k in _START_KEYWORDS):
                start_idx = col_idx
            if end_idx is None and any(k in text or k in low for k in _END_KEYWORDS):
                end_idx = col_idx
        if start_idx and end_idx:
            return row_idx, start_idx, end_idx, texts
    return None


def _hierarchy_columns(ws, header_row_idx: int, start_col_idx: int, header_texts: dict) -> list:
    """날짜 컬럼보다 왼쪽에 있는, 실제로 데이터가 들어있는 계층(파트/소분류/작업명) 컬럼들을
    좌->우 순서로 반환한다. 'No'류 순번 컬럼은 제외한다."""
    candidates = []
    for col_idx in range(1, start_col_idx):
        header = header_texts.get(col_idx, "")
        if _INDEX_COL_RE.match(header):
            continue
        has_data = any(
            ws.cell(row=r, column=col_idx).value not in (None, "")
            for r in range(header_row_idx + 1, ws.max_row + 1)
        )
        if has_data:
            candidates.append(col_idx)
    return candidates


def _excel_table_to_dict(ws) -> dict | None:
    header = _find_header_row(ws)
    if header is None:
        return None
    header_row_idx, start_col_idx, end_col_idx, header_texts = header

    hier_cols = _hierarchy_columns(ws, header_row_idx, start_col_idx, header_texts)
    if not hier_cols:
        return None

    rightmost = hier_cols[-1]
    # 2개 컬럼(파트+작업)뿐이면 가운데 소분류 컬럼이 따로 없으므로, subgroup은
    # task명으로 대체하는 기존 규칙을 쓴다 (leftmost와 겹치는 컬럼을 또 쓰지 않음).
    second_rightmost = hier_cols[-2] if len(hier_cols) >= 3 else None
    leftmost = hier_cols[0]

    # 템플릿 종류 감지: "롤업/요약 행"(하위 항목 없이 분류명만 있는 행)이 하나라도 있으면
    # 그 롤업 행들로만 파트 경계를 정하는 병합 셀 스타일 템플릿이다.
    # 롤업 행이 전혀 없으면(모든 행이 leaf) 매 행 leftmost 컬럼값을 그대로 파트명으로 쓰는
    # 단순 플랫 테이블 스타일이다. (leaf 행에서 leftmost가 부수적인 라벨로 또 쓰이는
    # 병합 셀 스타일과 구분하기 위해 필요)
    has_rollup_rows = False
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        rightmost_blank = ws.cell(row=row_idx, column=rightmost).value in (None, "")
        if rightmost_blank and any(
            ws.cell(row=row_idx, column=c).value not in (None, "") for c in hier_cols if c != rightmost
        ):
            has_rollup_rows = True
            break

    parts: list = []
    part_by_name: dict = {}
    current_part_name = ws.title
    current_subgroup_name = None

    def get_part(name: str) -> dict:
        if name not in part_by_name:
            part = {"name": name, "color": None, "subgroups": []}
            part_by_name[name] = part
            parts.append(part)
        return part_by_name[name]

    def get_subgroup(part: dict, name: str) -> dict:
        for sg in part["subgroups"]:
            if sg["name"] == name:
                return sg
        sg = {"name": name, "tasks": []}
        part["subgroups"].append(sg)
        return sg

    undated: list = []  # 날짜가 비어 있던 task: (part_name, subgroup_name, task_name) - 나중에 기본일자로 보정

    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        raw_values = {c: ws.cell(row=row_idx, column=c).value for c in hier_cols}
        rightmost_raw = raw_values.get(rightmost)
        rightmost_blank = rightmost_raw in (None, "")

        other_literal = [
            (c, raw_values[c]) for c in hier_cols if c != rightmost and raw_values.get(c) not in (None, "")
        ]

        if rightmost_blank:
            if has_rollup_rows and other_literal:
                # 하위 항목이 없는 롤업/요약 행 -> 새 파트 경계로 본다
                current_part_name = str(other_literal[0][1]).strip()
                current_subgroup_name = None
            continue

        if not has_rollup_rows and raw_values.get(leftmost) not in (None, ""):
            # 플랫 테이블 스타일: leaf 행마다 leftmost 컬럼이 곧 파트명
            current_part_name = str(raw_values[leftmost]).strip()
            current_subgroup_name = None

        # 이 행에 다른 계층 컬럼 값이 새로 있으면 소분류 갱신
        if second_rightmost is not None and raw_values.get(second_rightmost) not in (None, ""):
            current_subgroup_name = str(raw_values[second_rightmost]).strip()

        task_name = str(rightmost_raw).strip()
        if not task_name:
            continue

        start_str = _cell_to_date_str(ws.cell(row=row_idx, column=start_col_idx).value)
        end_str = _cell_to_date_str(ws.cell(row=row_idx, column=end_col_idx).value)
        sg_name = current_subgroup_name or task_name

        if not start_str or not end_str:
            # 날짜가 비어 있는 항목도 누락시키지 않고 기본일자로 보정해 포함한다
            undated.append((current_part_name, sg_name, task_name))
            continue

        part = get_part(current_part_name)
        sg = get_subgroup(part, sg_name)
        sg["tasks"].append({"name": task_name, "start_date": start_str, "end_date": end_str})

    all_dates = [t["start_date"] for p in parts for sg in p["subgroups"] for t in sg["tasks"]]
    all_dates += [t["end_date"] for p in parts for sg in p["subgroups"] for t in sg["tasks"]]
    if not all_dates and not undated:
        return None

    fallback_date = min(all_dates) if all_dates else _dt.date.today().isoformat()
    for part_name, sg_name, task_name in undated:
        part = get_part(part_name)
        sg = get_subgroup(part, sg_name)
        sg["tasks"].append({"name": task_name, "start_date": fallback_date, "end_date": fallback_date})

    all_dates = [t["start_date"] for p in parts for sg in p["subgroups"] for t in sg["tasks"]]
    all_dates += [t["end_date"] for p in parts for sg in p["subgroups"] for t in sg["tasks"]]

    return {
        "project_start": min(all_dates),
        "project_end": max(all_dates),
        "parts": parts,
        "phases": [],
    }


def _trim_trailing_empty(cells: list) -> list:
    while cells and not cells[-1].strip():
        cells.pop()
    return cells


def _excel_to_text(path: str) -> str:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = _select_wbs_sheet(wb)
    ws = wb[sheet_name]

    lines = [f"[시트: {ws.title}]"]
    for row in ws.iter_rows(values_only=True):
        cells = [str(c) if c is not None else "" for c in row]
        # 주간 그리드/숨김 헬퍼 컬럼처럼 0/1 같은 한 글자짜리 값만 길게 반복되는
        # 꼬리 컬럼은 노이즈만 되고 파싱에 쓸모가 없어 제거한다 (날짜는 이미 명시 컬럼에 있음).
        while len(cells) > 1 and len(cells[-1].strip()) <= 1:
            cells.pop()
        cells = _trim_trailing_empty(cells)
        if any(cell.strip() for cell in cells):
            lines.append("\t".join(cells))
    return "\n".join(lines)


def _pdf_to_text(path: str) -> str:
    from pypdf import PdfReader

    reader = PdfReader(path)
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _json_to_text(path: str) -> str:
    with open(path, "r", encoding="utf-8") as f:
        return f.read()


def _image_media_type(path: str) -> str:
    ext = os.path.splitext(path)[1].lower()
    if ext in (".jpg", ".jpeg"):
        return "image/jpeg"
    return "image/png"


# ---- draw.io 입력 (LLM 호출 없이 좌표 기반으로 직접 복원) ----
#
# 우리 렌더러(renderer.py)가 만드는 draw.io와 동일한 레이아웃 규칙
# (라벨영역 너비, 1주 간격 px, 날짜 헤더가 매주 월요일)을 가진 drawio 파일이라면
# 셀의 x/y/width/height 좌표만으로 파트>소분류>작업 계층과 날짜를 그대로 복원할 수 있다.
# 날짜 헤더 2개 이상의 x좌표 간격으로 그리드 스케일을 그 파일 기준으로 스스로 추정하므로
# 우리 config 상수와 다른 px 스케일을 쓴 drawio라도 동작한다.

import datetime as _dt
import re as _re
import xml.etree.ElementTree as _ET

_DATE_HEADER_RE = _re.compile(r"^\d{1,2}/\d{1,2}$")


def _drawio_style_dict(style: str) -> dict:
    d = {}
    for part in (style or "").split(";"):
        if "=" in part:
            k, v = part.split("=", 1)
            d[k] = v
    return d


def _drawio_base_color(fill: str) -> str:
    if fill and fill.startswith("#") and len(fill) > 7:
        return fill[:7]
    return fill or config.DEFAULT_PALETTE[0]


def _drawio_load_items(path: str) -> list:
    tree = _ET.parse(path)
    items = []
    for c in tree.getroot().findall(".//mxCell"):
        if c.get("vertex") != "1":
            continue
        geo = c.find("mxGeometry")
        if geo is None:
            continue
        st = _drawio_style_dict(c.get("style", ""))
        items.append(
            {
                "x": float(geo.get("x", 0)),
                "y": float(geo.get("y", 0)),
                "w": float(geo.get("width", 0)),
                "h": float(geo.get("height", 0)),
                "val": (c.get("value") or "").strip(),
                "fill": st.get("fillColor", ""),
            }
        )
    return items


def _drawio_calibrate(items: list):
    date_cells = sorted((it for it in items if _DATE_HEADER_RE.match(it["val"])), key=lambda i: i["x"])
    if len(date_cells) < 2:
        raise WBSParseError("draw.io 파일에서 날짜 헤더(예: 3/2, 3/9 ...)를 찾을 수 없습니다.")

    label_width = date_cells[0]["x"]
    week_width = date_cells[1]["x"] - date_cells[0]["x"]
    header_bottom = max(it["y"] + it["h"] for it in date_cells)
    month, day = map(int, date_cells[0]["val"].split("/"))

    now_year = _dt.date.today().year
    chosen_year = now_year
    for y in range(now_year - 1, now_year + 3):
        try:
            if _dt.date(y, month, day).weekday() == 0:  # 월요일
                chosen_year = y
                break
        except ValueError:
            continue

    grid_start = _dt.date(chosen_year, month, day)
    return label_width, week_width, grid_start, header_bottom


def _drawio_to_dict(path: str) -> dict:
    items = _drawio_load_items(path)
    label_width, week_width, grid_start, header_bottom = _drawio_calibrate(items)

    def x_to_date(x: float) -> _dt.date:
        days = (x - label_width) / week_width * 7
        return grid_start + _dt.timedelta(days=round(days))

    body_items = [it for it in items if it["y"] >= header_bottom - 1]

    label_rows = sorted(
        (
            it
            for it in body_items
            if 0 <= it["x"] <= label_width * 0.25 and it["w"] >= label_width * 0.5 and it["val"]
        ),
        key=lambda i: i["y"],
    )
    bars = [it for it in body_items if it["x"] >= label_width - 5 and 8 <= it["h"] <= 20]
    phase_rows = [
        it
        for it in items
        if it["y"] < 1 and it["x"] >= label_width - 1 and it["val"] and it["fill"] not in ("", "none")
    ]

    def find_bar(row: dict):
        for b in bars:
            if b["y"] >= row["y"] - 1 and b["y"] + b["h"] <= row["y"] + row["h"] + 1:
                return b
        return None

    parts: list = []
    cur_part = None
    cur_sub = None

    def ensure_part():
        nonlocal cur_part
        if cur_part is None:
            cur_part = {"name": "(미분류)", "color": config.DEFAULT_PALETTE[0], "subgroups": []}
            parts.append(cur_part)
        return cur_part

    def ensure_sub(default_name: str):
        nonlocal cur_sub
        if cur_sub is None:
            cur_sub = {"name": default_name, "tasks": []}
            ensure_part()["subgroups"].append(cur_sub)
        return cur_sub

    for row in label_rows:
        fill = (row["fill"] or "").lower()
        is_subgroup_bg = fill == "#f0f0f0"
        is_white_bg = fill in ("#ffffff", "#fff", "none", "")

        if 24 <= row["h"] < 28 and not is_subgroup_bg and not is_white_bg:
            cur_part = {"name": row["val"], "color": _drawio_base_color(row["fill"]), "subgroups": []}
            parts.append(cur_part)
            cur_sub = None
        elif 24 <= row["h"] < 28 and is_subgroup_bg:
            cur_sub = {"name": row["val"], "tasks": []}
            ensure_part()["subgroups"].append(cur_sub)
        elif row["h"] >= 28:
            bar = find_bar(row)
            if bar:
                start_date, end_date = x_to_date(bar["x"]), x_to_date(bar["x"] + bar["w"])
            else:
                start_date = end_date = grid_start
            sub = ensure_sub(row["val"])
            sub["tasks"].append(
                {"name": row["val"], "start_date": start_date.isoformat(), "end_date": end_date.isoformat()}
            )

    all_dates = [t["start_date"] for p in parts for sg in p["subgroups"] for t in sg["tasks"]]
    all_dates += [t["end_date"] for p in parts for sg in p["subgroups"] for t in sg["tasks"]]
    if not all_dates:
        raise WBSParseError("draw.io 파일에서 작업(task) 막대를 찾지 못했습니다.")

    phases = [
        {
            "name": p["val"],
            "color": _drawio_base_color(p["fill"]),
            "start_date": x_to_date(p["x"]).isoformat(),
            "end_date": x_to_date(p["x"] + p["w"]).isoformat(),
        }
        for p in sorted(phase_rows, key=lambda i: i["x"])
    ]

    return {
        "project_start": min(all_dates),
        "project_end": max(all_dates),
        "parts": parts,
        "phases": phases,
    }


def _parse_excel(path: str) -> dict:
    """엑셀이 명시적 시작/종료 날짜 컬럼을 가진 표 형태면 LLM 없이 좌표 기반으로
    정확하게 추출한다 (완전성 보장). 구조가 표준적이지 않아 실패하면
    텍스트 덤프 + LLM 파싱으로 폴백한다."""
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = _select_wbs_sheet(wb)
    ws = wb[sheet_name]

    result = _excel_table_to_dict(ws)
    if result is not None and sum(len(sg["tasks"]) for p in result["parts"] for sg in p["subgroups"]) > 0:
        return result

    text = _excel_to_text(path)
    return _call_llm_text(text)


def parse_wbs_file(path: str) -> dict:
    """입력 파일 경로를 받아 WBS 구조 dict를 반환한다.
    draw.io는 좌표 기반으로, 표 구조가 명확한 엑셀도 좌표 기반으로 LLM 호출 없이
    직접 복원한다. 그 외(JSON/PDF/이미지, 또는 구조가 불명확한 엑셀)는 선택된
    LLM 프로바이더로 파싱한다."""
    ext = os.path.splitext(path)[1].lower()

    if ext in config.SUPPORTED_DRAWIO_EXT:
        result = _drawio_to_dict(path)
    elif ext in config.SUPPORTED_IMAGE_EXT:
        with open(path, "rb") as f:
            image_bytes = f.read()
        result = _call_llm_vision(image_bytes, _image_media_type(path))
    elif ext in config.SUPPORTED_EXCEL_EXT:
        result = _parse_excel(path)
    elif ext in config.SUPPORTED_PDF_EXT:
        text = _pdf_to_text(path)
        result = _call_llm_text(text)
    elif ext in config.SUPPORTED_JSON_EXT:
        text = _json_to_text(path)
        result = _call_llm_text(text)
    else:
        raise WBSParseError(f"지원하지 않는 파일 형식입니다: {ext}")

    _assign_missing_colors(result)
    _sanitize_dates(result)
    return result


def _assign_missing_colors(result: dict) -> None:
    palette = config.DEFAULT_PALETTE
    idx = 0
    for part in result.get("parts", []):
        if not part.get("color"):
            part["color"] = palette[idx % len(palette)]
        idx += 1


def _fix_date_str(value: str, fallback: str) -> str:
    """LLM(특히 Vision)이 가끔 '2026-02-30'처럼 존재하지 않는 날짜를 반환하는 경우,
    해당 월의 마지막 유효한 날로 보정한다. 형식 자체가 깨졌으면 fallback을 사용한다."""
    import calendar
    import re

    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", str(value).strip())
    if not m:
        return fallback
    year, month, day = (int(g) for g in m.groups())
    if not (1 <= month <= 12):
        return fallback
    last_day = calendar.monthrange(year, month)[1]
    day = min(max(day, 1), last_day)
    return f"{year:04d}-{month:02d}-{day:02d}"


def _sanitize_dates(result: dict) -> None:
    fallback = result.get("project_start") or "2026-01-01"
    for part in result.get("parts", []):
        for sg in part.get("subgroups", []):
            for task in sg.get("tasks", []):
                task["start_date"] = _fix_date_str(task.get("start_date"), fallback)
                task["end_date"] = _fix_date_str(task.get("end_date"), task["start_date"])
                if task["end_date"] < task["start_date"]:
                    task["end_date"] = task["start_date"]
    for phase in result.get("phases", []):
        phase["start_date"] = _fix_date_str(phase.get("start_date"), fallback)
        phase["end_date"] = _fix_date_str(phase.get("end_date"), phase["start_date"])

    all_dates = [
        d
        for part in result.get("parts", [])
        for sg in part.get("subgroups", [])
        for task in sg.get("tasks", [])
        for d in (task["start_date"], task["end_date"])
    ]
    if all_dates:
        result["project_start"] = min(all_dates)
        result["project_end"] = max(all_dates)
